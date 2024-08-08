import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def auto_adjust_column_widths(excel_file_path: str, extra_space: int = 1) -> None:
    """
    Adjusts column widths of the excel file and replaces it with the adjusted one.
    Adjusting columns is based on the lengths of columns values (including column names).

    Parameters
    ----------
    excel_file_path : str
        Path to the excel file to adjust column widths.

    extra_space : int
        Extra column width in addition to the value-based widths.
    """
    wb = load_workbook(excel_file_path)

    for ws in wb.worksheets:
        # Create DataFrame including the header
        df = pd.DataFrame(ws.values)

        # Calculate maximum lengths of each column
        max_lengths = df.astype(str).apply(lambda col: col.str.len()).max(axis=0) + extra_space

        # Adjust column widths
        for i, r in max_lengths.items():
            ws.column_dimensions[get_column_letter(i + 1)].width = r

    wb.save(excel_file_path)

def format_YYYYMMDD(dt: datetime) -> str:
    return dt.strftime('%Y%m%d')
