import subprocess
import numpy as np
import pandas as pd
import os
from loguru import logger as L
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from libs import auto_adjust_column_widths, format_YYYYMMDD
import pytz
from export_edl_funds import fp
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.utils.dataframe import dataframe_to_rows
load_dotenv()

excel_file_path = fp
export_edl_funds_path = os.getenv('Export_Edl_Funds')
if not os.path.exists(excel_file_path):
    print(f"{excel_file_path} does not exist. Running {export_edl_funds_path} to generate the file")
    try:
        subprocess.run(['python', export_edl_funds_path], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Failed to run {export_edl_funds_path}: {e}")
        exit(1)

Root_Folder_Path = os.getenv("Root_Folder_Path")
L.info(f"Root_Folder_Path: {Root_Folder_Path}")
eFinance_Master_List_File_Name = os.getenv("eFinance_Master_List_File_Name")
eFinance_Master_List_Full_Path = os.path.join(Root_Folder_Path, eFinance_Master_List_File_Name)
L.info(f"eFinance_Master_List_Full_Path: {eFinance_Master_List_Full_Path}")

EDL_Master_List_FileName = os.getenv("EDL_Master_List_FileName")
EDL_Master_List_Output_Folder = os.getenv("EDL_Master_List_Output_Folder")
EDL_Master_List_Full_Path = os.path.join(Root_Folder_Path, EDL_Master_List_Output_Folder, EDL_Master_List_FileName)
L.info(f"EDL_Master_List_Full_Path: {EDL_Master_List_Full_Path}")

now_str = format_YYYYMMDD(datetime.now())
Fund_Code_Mapping_FilePath = os.path.join(Root_Folder_Path, EDL_Master_List_Output_Folder, f"Fund_Code_Mapping_Compare_EDL_eFinance_{now_str}.xlsx")
L.info("Fund_Code_Mapping_FilePath", Fund_Code_Mapping_FilePath)
# print(Fund_Code_Mapping_FilePath)

df_raw_eFinance_Master = pd.read_excel(eFinance_Master_List_Full_Path, sheet_name='Data')

# eFinance_exclude_platform_code_list = [
# 'BB_GA',
# 'BB_GA_LIFECO',
# 'CH_GA',
# 'CH_INST',
# 'CH_WOFE',
# 'HK_ADV',
# 'HK_GA_LIFECO',
# 'HK_GA_INTCOLOAN',
# 'HK_GA_MFIL',
# 'HK_GA_MLRL',
# 'HK_GA_RE',
# 'HK_INST',
# 'HK_MPF_APIF1',
# 'HK_MPF_APIF2',
# 'HK_MPF_MACAU',
# 'HK_MPF_ORSO',
# 'HK_MPF_SC',
# 'ID_DPLK',
# 'ID_GA_LIFECO',
# 'ID_ILP',
# 'ID_INST',
# 'JP_GA',
# 'JP_GA_LIFECO',
# 'JP_GA_RE',
# 'JP_LOCAL',
# 'JP_PRIVATE',
# 'KH_GA_LIFECO',
# 'MY_3P',
# 'MY_GA',
# 'MY_INST',
# 'MY_PRS',
# 'OT_TBD',
# 'PH_GA',
# 'PH_ILP',
# 'PH_ILP_MCBL',
# 'PH_ILP_MP',
# 'PH_INST',
# 'SG_GA_LIFECO',
# 'SG_GA_RE',
# 'SG_ILP',
# 'SG_INST',
# 'SG_INST_IR_ICAV',
# 'SG_MAF',
# 'SG_MGF',
# 'SG_OTHER',
# 'SG_REIT',
# 'TH_GA_LIFECO',
# 'TH_MAM',
# 'TW_GA_LIFECO',
# 'VN_GA_LIFECO',
# 'VN_ILP',
# 'VN_MAM',
# 'VN_OTHER',
# np.nan
# ]
eFinance_exclude_TA_Scope_list = [
    'BASE',
    'BB_GA',
    'HK_MIT',
    'IN_MAHINDRA',
    'MY_INST',
    'VN_ILP',
    #
    'CH_INST',
    'CH_WOFE',
    'HK_ADV',
    'HK_EAGL',
    'HK_INST',
    'HK_SCB_CASH',
    'ID_CAS4TA_PRICES',
    'ID_DPLK',
    'ID_ILP',
    'ID_INST',
    'JP_GA',
    'JP_GX',
    'JP_INST',
    'JP_LIAISON',
    'JP_NRI',
    'JP_PRIVATE',
    'JP_SUBADVISORY',
    'MY_ILP',
    'MY_PRS',
    'OT_OTHERFUNDS',
    'PH_GA',
    'PH_ILP',
    'PH_INST',
    'SG_GA',
    'SG_ILP',
    'SG_INST',
    'SG_INST_IR_ICAV',
    'SG_MAF',
    'SG_MGF',
    'SG_OTHER',
    'SG_REIT',
    'TH_MAM',
    'VN_GA',
    'VN_MAM',
    'VN_OTHER',
    np.nan
]
# Filter out eFinance TA scopes

# df_raw_eFinance_Master['efinance_unique_key'] = df_raw_eFinance_Master['ISO2CountryCode'] + '_' + df_raw_eFinance_Master['local_ta_code']

df_eFin = df_raw_eFinance_Master[~df_raw_eFinance_Master['TAScopeCode'].isin(eFinance_exclude_TA_Scope_list) & ~((df_raw_eFinance_Master['ISO2Code'].str.contains('CN')) & (df_raw_eFinance_Master['ShareClassInvestorType'] == 'Institutional'))]

databricks_df = pd.read_excel(EDL_Master_List_Full_Path)

merged_df_1 = pd.merge(df_eFin[df_eFin['efin_isin_code'].notnull()], 
                       databricks_df[databricks_df['edl_market_id'].notnull()], 
                       left_on='efin_isin_code', right_on='edl_market_id', 
                       how='left')
merged_df_2 = pd.merge(df_eFin[df_eFin['efin_isin_code'].notnull()], 
                       databricks_df[databricks_df['edl_market_id'].isnull()],
                       left_on='efin_edl_align_unique_key', right_on='edl_unique_key', 
                       how='left')
merged_df_3 = pd.merge(df_eFin[df_eFin['efin_isin_code'].isnull()], 
                       databricks_df, 
                       left_on='efin_edl_align_unique_key', right_on='edl_unique_key', 
                       how='left')
merged_df = pd.concat([merged_df_1, merged_df_2, merged_df_3], ignore_index=True)

mapping_df = merged_df[['efin_edl_align_unique_key', 'edl_unique_key', 'TAScopeCode', 'efin_edl_align_ta_code', 'edl_fund_class_code_1', 'LastProvideDate', 'Year2023Flag', 'Name', 'edl_liability_portf_name', 'efin_isin_code', 'edl_market_id']]

unique_values = mapping_df.drop_duplicates()

databricks_df.to_excel(Fund_Code_Mapping_FilePath, sheet_name='EDL_Funds', index=False)

with pd.ExcelWriter(Fund_Code_Mapping_FilePath, engine='openpyxl', mode='a') as writer:
    df_eFin.to_excel(writer, sheet_name='eFinance_Funds', index=False)
    mapping_df.to_excel(writer, sheet_name='Compare', index=False)

wb = load_workbook(Fund_Code_Mapping_FilePath)
ws = wb['Compare']

wb.save(Fund_Code_Mapping_FilePath)

auto_adjust_column_widths(Fund_Code_Mapping_FilePath)
L.info(f"Done. {Fund_Code_Mapping_FilePath}")

# wb = load_workbook(Fund_Code_Mapping_FilePath)
# df = pd.read_excel(Fund_Code_Mapping_FilePath, sheet_name = 'Compare')

# pivot_columns = ['efin_edl_align_unique_key', 'edl_unique_key', 'TAScopeCode', 
#                  'efin_edl_align_ta_code', 'edl_fund_class_code_1', 'LastProvideDate', 
#                  'Year2023Flag']

# pivot_df = df[pivot_columns]

# pivot_sheet = wb.create_sheet(title='Pivot_Compare')

# for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index = False, header = True), 1):
#     for c_idx, value in enumerate(row, 1):
#         pivot_sheet.cell(row=r_idx, column=c_idx, value = value)


# pivot_table = Table(displayName = 'PivotTable1', ref=f'A1:G{len(pivot_df)+1}')
# style = TableStyleInfo(
#     name = 'TableStyleMedium9',
#     showFirstColumn=False,
#     showLastColumn=False,
#     showRowStripes=False,
#     showColumnStripes=False
# )
# pivot_table.tableStyleInfo = style
# pivot_sheet.add_table(pivot_table)
# wb.save(Fund_Code_Mapping_FilePath)
